"""Asserted against the pure builder rather than through the page.

`tests/test_data_cleaner_page.py` already records that `AppTest` cannot read the bytes
behind a download button, so a page-level test could only check that the button exists —
which `test_dashboard_page.py` does. What the workbook actually contains is checked here.
"""

import base64
import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from dashboard import images
from dashboard.exceptions import ReportExportError
from dashboard.excel_export import CONTENTS_SHEET_NAME, build_report_workbook, sheet_names_for
from dashboard.model import PinnedItem, Report, add_section, add_subsection, assign_item

# A real 1×1 PNG rather than a few token bytes: xlsxwriter reads the header to size the
# picture, so a fake would fail inside the writer for reasons that say nothing about this
# code.
FAKE_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
)


@pytest.fixture(autouse=True)
def no_real_rasterizing(monkeypatch):
    """Rasterizing launches a headless browser; the workbook only needs *some* bytes."""
    monkeypatch.setattr(images, "figure_to_png", lambda figure, **kwargs: FAKE_PNG)


@pytest.fixture
def frame() -> pd.DataFrame:
    return pd.DataFrame({"region": ["North", "South"], "sales": [120, 340]})


def _one_item_report(frame: pd.DataFrame, **item_fields) -> Report:
    report = Report(title="Q3 review")
    section = add_section(report, "Sales")
    item = PinnedItem(heading="Sales by region", frame=frame, **item_fields)
    report.pool.append(item)
    assign_item(report, item.item_id, section.subsections[0].node_id)
    return report


def _sheets(workbook_bytes: bytes) -> list[str]:
    return load_workbook(io.BytesIO(workbook_bytes)).sheetnames


# --------------------------------------------------------------------------------------
# Sheet naming
# --------------------------------------------------------------------------------------


def test_one_sheet_per_subsection_plus_contents(frame):
    report = Report(title="Two parts")
    section = add_section(report, "Sales")
    add_subsection(section, "By product")
    for subsection in section.subsections:
        item = PinnedItem(frame=frame)
        report.pool.append(item)
        assign_item(report, item.item_id, subsection.node_id)

    assert _sheets(build_report_workbook(report)) == [CONTENTS_SHEET_NAME, "1.1 General", "1.2 By product"]


def test_long_subsection_names_are_truncated_for_excel(frame):
    report = _one_item_report(frame)
    report.sections[0].subsections[0].name = "A" * 60

    name = _sheets(build_report_workbook(report))[1]
    assert len(name) <= 31


def test_forbidden_characters_are_replaced(frame):
    report = _one_item_report(frame)
    report.sections[0].subsections[0].name = "North/South: [2024]"

    assert "/" not in _sheets(build_report_workbook(report))[1]


def test_duplicate_subsection_names_are_still_distinct_sheets(frame):
    report = Report(title="Duplicates")
    first = add_section(report, "Sales")
    second = add_section(report, "Costs")
    for section in (first, second):
        section.subsections[0].name = "Summary"
        item = PinnedItem(frame=frame)
        report.pool.append(item)
        assign_item(report, item.item_id, section.subsections[0].node_id)

    sheets = _sheets(build_report_workbook(report))
    assert len(set(sheets)) == len(sheets)


def test_a_subsection_called_contents_does_not_displace_the_contents_sheet(frame):
    report = _one_item_report(frame)
    report.sections[0].subsections[0].name = ""

    names = sheet_names_for(report)
    assert CONTENTS_SHEET_NAME not in names.values()


def test_empty_subsections_get_no_sheet(frame):
    report = _one_item_report(frame)
    add_subsection(report.sections[0], "Nothing here")

    assert _sheets(build_report_workbook(report)) == [CONTENTS_SHEET_NAME, "1.1 General"]


# --------------------------------------------------------------------------------------
# Contents
# --------------------------------------------------------------------------------------


def test_the_contents_sheet_lists_the_title_and_every_subsection(frame):
    report = _one_item_report(frame)
    sheet = load_workbook(io.BytesIO(build_report_workbook(report)))[CONTENTS_SHEET_NAME]
    text = "\n".join(str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None)

    assert "Q3 review" in text
    assert "1. Sales" in text
    assert "1.1 General" in text


# --------------------------------------------------------------------------------------
# Sheet contents
# --------------------------------------------------------------------------------------


def test_a_sheet_carries_the_heading_the_comment_and_every_row(frame):
    report = _one_item_report(frame, comment="Sales held up in the North.")
    sheet = load_workbook(io.BytesIO(build_report_workbook(report)))["1.1 General"]
    text = "\n".join(str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None)

    assert "Sales by region" in text
    assert "Sales held up in the North." in text
    assert "North" in text and "South" in text
    assert "120" in text and "340" in text


def test_a_full_table_is_written_with_no_row_limit():
    frame = pd.DataFrame({"n": range(500)})
    report = _one_item_report(frame)
    sheet = load_workbook(io.BytesIO(build_report_workbook(report)))["1.1 General"]

    values = {cell.value for row in sheet.iter_rows() for cell in row}
    assert 0 in values and 499 in values


def test_a_chart_is_embedded_as_a_picture(frame):
    report = _one_item_report(frame, figure=object())
    sheet = load_workbook(io.BytesIO(build_report_workbook(report)))["1.1 General"]

    assert len(sheet._images) == 1


def test_a_chart_that_could_not_be_drawn_leaves_a_note_instead(monkeypatch, frame):
    monkeypatch.setattr(images, "figure_to_png", lambda figure, **kwargs: None)
    report = _one_item_report(frame, figure=object())
    sheet = load_workbook(io.BytesIO(build_report_workbook(report)))["1.1 General"]
    text = "\n".join(str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None)

    assert "couldn't be included as a picture" in text
    assert "North" in text


# --------------------------------------------------------------------------------------
# Refusals
# --------------------------------------------------------------------------------------


def test_an_empty_report_is_refused_with_a_readable_message():
    with pytest.raises(ReportExportError, match="place at least one"):
        build_report_workbook(Report(title="Nothing"))


def test_a_report_with_only_empty_sections_is_refused():
    report = Report(title="Skeleton")
    add_section(report, "Sales")
    with pytest.raises(ReportExportError):
        build_report_workbook(report)


def test_a_cell_over_excels_length_limit_is_refused_before_writing():
    frame = pd.DataFrame({"note": ["x" * 40_000]})
    with pytest.raises(ReportExportError, match="characters"):
        build_report_workbook(_one_item_report(frame))
